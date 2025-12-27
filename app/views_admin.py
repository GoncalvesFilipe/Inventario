from openpyxl import load_workbook, Workbook
from django.conf import settings
import os
import json
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .forms import PatrimonioForm, InventarianteUserForm
from .models import Inventariante, Patrimonio
from .decorators import admin_required
from .decorators import presidente_ou_superuser


# ==========================================================
# DASHBOARD ADMINISTRATIVO
# Fun√ß√£o respons√°vel por apresentar a p√°gina inicial do painel
# administrativo, acess√≠vel apenas mediante autentica√ß√£o.
# ==========================================================
@login_required
def admin_dashboard(request):
    return render(request, "app_inventario/admin_dashboard.html")


# ==========================================================
# FUN√á√ÉO DE AUTORIZA√á√ÉO
# ----------------------------------------------------------
# Respons√°vel por centralizar a regra de autoriza√ß√£o aplicada
# √†s rotinas administrativas sens√≠veis do sistema.
#
# O acesso √© concedido exclusivamente a:
# - Superusu√°rio
# - Inventariante com flag 'presidente=True'
# ==========================================================
def presidente_ou_superuser(user):
    return (
        user.is_authenticated and (
            user.is_superuser or
            (hasattr(user, "inventariante") and user.inventariante.presidente)
        )
    )


# ==========================================================
# FECHAMENTO DE MODAL
# Retorna uma resposta vazia utilizada pelo HTMX para encerrar
# modais sem renderizar conte√∫do adicional.
# ==========================================================
def close_modal(request):
    return HttpResponse("")


# ==========================================================
# LISTAGEM DE INVENTARIANTES
# Apresenta a lista completa de inventariantes cadastrados.
# A consulta n√£o implementa filtragem e √© destinada ao painel.
# ==========================================================
@login_required
def inventariantes_list(request):
    inventariantes = Inventariante.objects.all()
    return render(
        request,
        "app_inventario/partials/inventariantes_list.html",
        {"inventariantes": inventariantes}
    )


# ==========================================================
# LISTA PARCIAL (PARA HTMX)
# Renderiza apenas o fragmento da lista, empregado em recargas
# ass√≠ncronas sem reprocessamento da p√°gina completa.
# ==========================================================
@login_required
def inventariante_list_partial(request):
    inventariantes = Inventariante.objects.all()
    return render(
        request,
        "app_inventario/partials/inventariante_list.html",
        {"inventariantes": inventariantes}
    )

# ==========================================================
# ADI√á√ÉO DE INVENTARIANTE
# Implementa o fluxo completo para cria√ß√£o de usu√°rio vinculado
# a um inventariante. O acesso √© restrito a administradores.
# Utiliza HTMX para atualiza√ß√£o din√¢mica da interface.
# ==========================================================
@admin_required
def inventariante_add(request):
    if request.method == "POST":
        form = InventarianteUserForm(request.POST)

        if form.is_valid():
            form.save()

            # Renderiza um HTML de sucesso dentro do modal
            html = render_to_string(
                "app_inventario/partials/inventariante_success.html",
                {"mensagem": "Usu√°rio inventariante salvo com sucesso!"},
                request=request
            )
            return HttpResponse(html)

        # Reapresenta o formul√°rio com erros no contexto do modal
        html = render_to_string(
            "app_inventario/partials/inventariante_form.html",
            {"form": form},
            request=request
        )
        return HttpResponse(html)

    # Solicita√ß√£o inicial (GET) ‚Üí formul√°rio limpo
    form = InventarianteUserForm()
    return render(
        request,
        "app_inventario/partials/inventariante_form.html",
        {"form": form}
    )


# ==========================================================
# EDI√á√ÉO DE INVENTARIANTE
# ----------------------------------------------------------
# Respons√°vel por atualizar os dados vinculados a um inventariante.
# ==========================================================
@admin_required
def inventariante_edit(request, pk):
    inventariante = get_object_or_404(Inventariante, pk=pk)
    user = inventariante.user

    if request.method == "POST":
        form = InventarianteUserForm(request.POST, instance=inventariante, user=user)

        if form.is_valid():
            form.save()

            # Renderiza mensagem de sucesso dentro do modal
            html = render_to_string(
                "app_inventario/partials/inventariante_edit_success.html",
                {"mensagem": "Usu√°rio inventariante atualizado com sucesso!"},
                request=request
            )

            response = HttpResponse(html)
            # Dispara trigger para atualizar a lista
            response["HX-Trigger"] = json.dumps({
                "reload": True
            })
            return response

        # Se houver erros, reapresenta o formul√°rio
        html = render_to_string(
            "app_inventario/partials/inventariante_form.html",
            {
                "form": form,
                "inventariante": inventariante,
                "is_edit": True   # <<< ESSENCIAL para diferenciar edi√ß√£o
            },
            request=request
        )
        return HttpResponse(html)

    # GET inicial ‚Üí formul√°rio preenchido
    form = InventarianteUserForm(instance=inventariante, user=user)
    return render(
        request,
        "app_inventario/partials/inventariante_form.html",
        {
            "form": form,
            "inventariante": inventariante,
            "is_edit": True   # <<< ESSENCIAL para diferenciar edi√ß√£o
        }
    )


# ==========================================================
# CONFIRMA√á√ÉO DE EXCLUS√ÉO DE INVENTARIANTE
# ----------------------------------------------------------
# Exibe modal solicitando confirma√ß√£o pr√©via antes de proceder
# √† remo√ß√£o definitiva do registro.
# ==========================================================
@admin_required
def inventariante_delete_confirm(request, pk):
    inventariante = get_object_or_404(Inventariante, pk=pk)
    return render(
        request,
        "app_inventario/partials/inventariante_delete_confirm.html",
        {"inventariante": inventariante}
    )


# ==========================================================
# EXCLUS√ÉO DE INVENTARIANTE
# ----------------------------------------------------------
# Realiza a remo√ß√£o tanto do registro Inventariante quanto do
# usu√°rio associado, caso exista, e aciona recarga din√¢mica da
# listagem via HTMX.
# ==========================================================
@admin_required
def inventariante_delete(request, pk):
    inventariante = get_object_or_404(Inventariante, pk=pk)

    if request.method == "POST":
        try:
            if inventariante.user:
                inventariante.user.delete()
        except Exception:
            pass

        inventariante.delete()

        inventariantes = Inventariante.objects.all()
        html = render_to_string(
            "app_inventario/partials/inventariantes_list.html",
            {"inventariantes": inventariantes},
            request=request
        )

        response = HttpResponse(html)
        response["HX-Trigger"] = json.dumps({
            "closeModal": True,
            "reloadInventariantes": True
        })
        return response

    return HttpResponse(status=405)



# ==========================================================
# PATRIM√îNIOS
# ----------------------------------------------------------
# M√≥dulo respons√°vel pelo gerenciamento completo dos bens
# patrimoniais, incluindo listagem, cadastro, edi√ß√£o e exclus√£o.
# ==========================================================

# ==========================================================
# LISTA DE PATRIM√îNIOS
# ----------------------------------------------------------
# Implementa busca, pagina√ß√£o e segrega√ß√£o dos resultados com
# base no perfil do usu√°rio (administrador ou inventariante).
# ==========================================================
@login_required
def patrimonio_list(request):
    # Recupera par√¢metros de busca e pagina√ß√£o
    search_query = request.GET.get('q')
    pagina_numero = request.GET.get('page', 1)

    # ------------------------------------------------------
    # Regra de permiss√£o unificada:
    # - Superusu√°rio
    # - Inventariante Presidente
    # ------------------------------------------------------
    is_admin = (
        request.user.is_superuser or
        (hasattr(request.user, "inventariante") and request.user.inventariante.presidente)
    )

    # ------------------------------------------------------
    # Defini√ß√£o do Queryset base
    # ------------------------------------------------------
    if is_admin:
        patrimonios = Patrimonio.objects.all()
    else:
        inventariante = get_object_or_404(Inventariante, user=request.user)
        patrimonios = Patrimonio.objects.filter(inventariante=inventariante)

    # ------------------------------------------------------
    # Filtro de busca textual
    # ------------------------------------------------------
    if search_query:
        patrimonios = patrimonios.filter(
            Q(tombo__icontains=search_query) |
            Q(descricao__icontains=search_query) |
            Q(setor__icontains=search_query) |
            Q(dependencia__icontains=search_query) |
            Q(fornecedor__icontains=search_query) |
            Q(numero_documento__icontains=search_query) |
            Q(conta_contabil__icontains=search_query)
        )

    # ------------------------------------------------------
    # Pagina√ß√£o (6 itens por p√°gina)
    # ------------------------------------------------------
    paginator = Paginator(patrimonios.order_by('id'), 6)
    try:
        lista_patrimonios = paginator.page(pagina_numero)
    except (PageNotAnInteger, EmptyPage):
        lista_patrimonios = paginator.page(1)

    # ------------------------------------------------------
    # Contexto unificado
    # ------------------------------------------------------
    context = {
        "pagina": "patrimonio",
        "lista_patrimonios": lista_patrimonios,
        "search_query": search_query or "",
        "is_admin": is_admin,
        "user": request.user,
    }
    
    # ------------------------------------------------------
    # L√≥gica HTMX:
    # - Se o alvo for apenas a tabela, renderiza parcial
    # - Caso contr√°rio, renderiza p√°gina completa
    # ------------------------------------------------------
    if request.headers.get('HX-Target') == 'conteudo-patrimonios':
        return render(request, "app_inventario/partials/tabela_patrimonios.html", context)
    
    return render(request, "app_inventario/patrimonio_list.html", context)


# ==========================================================
# FORMUL√ÅRIO HTMX DE PATRIM√îNIO
# ----------------------------------------------------------
# Permite inventariantes comuns adicionarem seus pr√≥prios bens.
# ==========================================================
@login_required
def patrimonio_form(request):
    inventariante = get_object_or_404(Inventariante, user=request.user)
    if request.method == "POST":
        form = PatrimonioForm(request.POST)
        if form.is_valid():
            patrimonio = form.save(commit=False)
            patrimonio.inventariante = inventariante
            patrimonio.save()
            tabela = render_to_string(
                "app_inventario/partials/tabela_patrimonios.html",
                {"patrimonios": Patrimonio.objects.filter(inventariante=inventariante)},
                request=request
            )
            return HttpResponse(tabela)
    return render(request, "app_inventario/partials/form_patrimonio.html", {"form": PatrimonioForm()})


# ==========================================================
# ADI√á√ÉO DE PATRIM√îNIO
# ----------------------------------------------------------
# Restrito a superusu√°rio ou inventariante presidente.
# ==========================================================
@login_required
@user_passes_test(presidente_ou_superuser)
def patrimonio_add(request):
    inventariante = get_object_or_404(Inventariante, user=request.user)
    if request.method == "POST":
        form = PatrimonioForm(request.POST, user=request.user)
        if form.is_valid():
            patrimonio = form.save(commit=False)
            patrimonio.inventariante = inventariante
            patrimonio.save()
            return HttpResponse("OK")
    return render(request, "app_inventario/partials/form_patrimonio.html", {"form": PatrimonioForm(user=request.user)})

# ==========================================================
# EDI√á√ÉO DE PATRIM√îNIO
# ----------------------------------------------------------
# Restrito a superusu√°rio ou inventariante presidente.
# ==========================================================
@login_required
@user_passes_test(presidente_ou_superuser)
def patrimonio_edit(request, pk):
    patrimonio = get_object_or_404(Patrimonio, pk=pk)
    if request.method == "POST":
        form = PatrimonioForm(request.POST, request.FILES, instance=patrimonio)
        if form.is_valid():
            form.save()
            return HttpResponse("", headers={"HX-Refresh": "true"})
    return render(request, "app_inventario/partials/form_patrimonio.html", {
        "form": PatrimonioForm(instance=patrimonio),
        "modo_edicao": True,
        "patrimonio": patrimonio
    })

# ==========================================================
# CONFIRMA√á√ÉO DE EXCLUS√ÉO DE PATRIM√îNIO
# ----------------------------------------------------------
# Restrito a superusu√°rio ou inventariante presidente.
# ==========================================================
@login_required
@user_passes_test(presidente_ou_superuser)
def confirmar_exclusao_patrimonio(request, pk):
    patrimonio = get_object_or_404(Patrimonio, pk=pk)
    return render(request, "app_inventario/partials/confirmar_exclusao_patrimonio.html", {"patrimonio": patrimonio})


# ==========================================================
# EXCLUS√ÉO DE PATRIM√îNIO
# ----------------------------------------------------------
# Restrito a superusu√°rio ou inventariante presidente.
# ==========================================================
@login_required
@user_passes_test(presidente_ou_superuser)
def excluir_patrimonio(request, pk):
    if request.method != "POST":
        return HttpResponse(status=405)
    
    patrimonio = get_object_or_404(Patrimonio, pk=pk)
    patrimonio.delete()
    
    # Recarrega a lista para atualizar a tabela via HTMX
    is_admin = presidente_ou_superuser(request.user)
    patrimonios_all = Patrimonio.objects.all().order_by('id')
    paginator = Paginator(patrimonios_all, 6)
    lista_patrimonios = paginator.page(1)

    context = {
        "lista_patrimonios": lista_patrimonios,
        "is_admin": is_admin,
    }

    # Renderiza apenas o fragmento da tabela
    html = render_to_string(
        "app_inventario/partials/tabela_patrimonios.html", 
        context, 
        request=request
    )
    
    response = HttpResponse(html)
    # Dispara os eventos para fechar o modal e avisar o frontend
    response["HX-Trigger"] = json.dumps({
        "patrimonioExcluido": True,
        "closeModal": True 
    })
    return response

# ==========================================================
# EXCLUS√ÉO DE PLANILHA DE PATRIM√îNIO
# ----------------------------------------------------------
# Respons√°vel por remover o arquivo f√≠sico da planilha Excel
# e tamb√©m excluir todos os registros de patrim√¥nio do banco.
#
# Fluxo:
# 1) Recebe requisi√ß√£o POST via HTMX
# 2) Remove o arquivo "registros.xlsx" do diret√≥rio MEDIA_ROOT
# 3) Exclui todos os registros da tabela Patrimonio
# 4) Dispara evento HTMX "planilhaExcluida" para feedback
#    visual e recarregamento da p√°gina
# ==========================================================
@login_required
@user_passes_test(presidente_ou_superuser)
def excluir_planilha(request):
    # ------------------------------------------------------
    # Valida√ß√£o do m√©todo HTTP
    # ------------------------------------------------------
    if request.method != "POST":
        return HttpResponse(status=405)

    # ------------------------------------------------------
    # Caminho absoluto da planilha
    # ------------------------------------------------------
    caminho_planilha = os.path.join(settings.MEDIA_ROOT, "registros.xlsx")

    # ------------------------------------------------------
    # Remo√ß√£o do arquivo f√≠sico, se existir
    # ------------------------------------------------------
    if os.path.exists(caminho_planilha):
        os.remove(caminho_planilha)

    # ------------------------------------------------------
    # Exclus√£o de todos os registros de patrim√¥nio no banco
    # ------------------------------------------------------
    Patrimonio.objects.all().delete()

    # ------------------------------------------------------
    # Retorno da resposta + disparo de evento HTMX
    # ------------------------------------------------------
    response = HttpResponse("")
    response["HX-Trigger"] = json.dumps({
        "planilhaExcluida": True
    })
    return response


# ==========================================================
# CONFIRMA√á√ÉO DE EXCLUS√ÉO DE PLANILHA
# ----------------------------------------------------------
# Exibe modal solicitando confirma√ß√£o pr√©via antes de proceder
# √† remo√ß√£o definitiva da planilha e dos registros vinculados.
# ==========================================================
@login_required
@user_passes_test(presidente_ou_superuser)
def excluir_planilha_confirm(request):
    return render(
        request,
        "app_inventario/partials/excluir_planilha_confirm.html"
    )



# ==========================================================
# REGISTRO EM PLANILHA
# ----------------------------------------------------------
# View respons√°vel por inserir dados em uma planilha Excel,
# associando explicitamente o registro ao inventariante
# autenticado e criando tamb√©m o registro no banco de dados.
# Esta funcionalidade √© restrita ao painel administrativo.
# ==========================================================
@login_required
def adicionar_na_planilha(request):
    """
    Registra uma nova linha em planilha Excel, criando o arquivo
    automaticamente caso ainda n√£o exista no diret√≥rio media.
    Al√©m disso, insere o registro no banco de dados Patrimonio.
    """

    #  Recupera usu√°rio e inventariante vinculado
    usuario = request.user
    inventariante = get_object_or_404(Inventariante, user=usuario)

    # Define caminho absoluto do arquivo Excel
    caminho_planilha = os.path.join(
        settings.MEDIA_ROOT,  # usar MEDIA_ROOT em vez de BASE_DIR
        "registros.xlsx"
    )

    # Verifica se a planilha j√° existe
    if os.path.exists(caminho_planilha):
        # Abre planilha existente
        workbook = load_workbook(caminho_planilha)
        sheet = workbook.active
    else:
        # Cria nova planilha
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registros"

        # Cabe√ßalho inicial
        sheet.append([
            "ID Usu√°rio",
            "Username",
            "Nome Completo",
            "Descri√ß√£o",
            "Valor"
        ])

    # Montagem da nova linha
    nova_linha = [
        usuario.id,
        usuario.username,
        usuario.get_full_name(),
        "Descri√ß√£o do registro",
        150.00
    ]

    # Inser√ß√£o e salvamento na planilha
    sheet.append(nova_linha)
    workbook.save(caminho_planilha)

    # Inser√ß√£o no banco de dados Patrimonio
    # Gera n√∫mero de patrim√¥nio automaticamente:
    # pega o √∫ltimo ID e soma 1
    ultimo = Patrimonio.objects.order_by("-id").first()
    if ultimo:
        patrimonio_numero = ultimo.patrimonio + 1
    else:
        patrimonio_numero = 1

    descricao = "Descri√ß√£o do registro"
    setor = "Setor padr√£o"
    dependencia = "Depend√™ncia padr√£o"
    situacao = "localizado"

    Patrimonio.objects.create(
        patrimonio=patrimonio_numero,
        descricao=descricao,
        setor=setor,
        dependencia=dependencia,
        situacao=situacao,
        inventariante=inventariante
    )

    # Retorno da resposta
    return HttpResponse("Registro salvo com sucesso")

# ==========================================================
# REGISTRO DE PLANILHA (UPLOAD MANUAL)
# ----------------------------------------------------------
# View respons√°vel exclusivamente pelo processamento 
# da planilha oficial de registros do sistema.
#
# Funcionalidade restrita a usu√°rios com perfil:
# - Presidente
# - Superusu√°rio
#
# Ap√≥s o processamento:
# - Os dados s√£o inseridos no banco de dados
# - A tabela de patrim√¥nios √© renderizada e devolvida
#   diretamente ao frontend via HTMX
# ==========================================================
import json
import os
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, user_passes_test
from openpyxl import load_workbook
from .models import Inventariante, Patrimonio
from .utils import presidente_ou_superuser  # supondo que voc√™ tenha esse helper

@login_required
@user_passes_test(presidente_ou_superuser)
def upload_planilha(request):
    if request.method != "POST":
        return HttpResponse(status=405)

    if not request.FILES.get("planilha"):
        return HttpResponse(status=400)

    planilha = request.FILES["planilha"]

    # ------------------------------------------------------
    # Garante diret√≥rio de m√≠dia e salva arquivo enviado
    # ------------------------------------------------------
    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
    caminho_planilha = os.path.join(settings.MEDIA_ROOT, "registros.xlsx")
    fs = FileSystemStorage(location=settings.MEDIA_ROOT)
    fs.save("registros.xlsx", planilha)

    # ------------------------------------------------------
    # Abre planilha com openpyxl
    # ------------------------------------------------------
    workbook = load_workbook(caminho_planilha)
    sheet = workbook.active

    # ------------------------------------------------------
    # Recupera inventariante vinculado ao usu√°rio logado
    # ------------------------------------------------------
    inventariante = get_object_or_404(Inventariante, user=request.user)

    # ------------------------------------------------------
    # Itera linhas a partir da segunda (linha 1 em branco)
    # ------------------------------------------------------
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):  # ignora linhas totalmente vazias
            continue

        # üîπ Leitura por √≠ndices fixos (colunas da planilha)
        patrimonio_numero = row[0]   # Coluna A ‚Üí Tombo
        descricao = row[1]           # Coluna B ‚Üí Descri√ß√£o
        valor = row[2]               # Coluna C ‚Üí Valor
        conta_contabil = row[3]      # Coluna D ‚Üí Conta Cont√°bil
        setor = row[4]               # Coluna E ‚Üí Setor
        empenho = row[5]             # Coluna F ‚Üí Empenho
        fornecedor = row[6]          # Coluna G ‚Üí Fornecedor
        numero_documento = row[7]    # Coluna H ‚Üí N¬∫ Documento
        data_documento = row[8]      # Coluna I ‚Üí Data Documento
        data_ateste = row[9]         # Coluna J ‚Üí Data Ateste
        dependencia = row[10]        # Coluna K ‚Üí Depend√™ncia

        print("Linha lida:", patrimonio_numero, descricao, valor, conta_contabil,
              setor, empenho, fornecedor, numero_documento, data_documento,
              data_ateste, dependencia)  # Debug

        # --------------------------------------------------
        # Valida n√∫mero do patrim√¥nio (tombo)
        # --------------------------------------------------
        try:
            patrimonio_numero = int(patrimonio_numero)
        except (TypeError, ValueError):
            continue

        # --------------------------------------------------
        # Cria registro no banco vinculado ao inventariante
        # --------------------------------------------------
        Patrimonio.objects.create(
            tombo=patrimonio_numero,
            descricao=descricao or "",
            valor=valor or None,
            conta_contabil=conta_contabil or "",
            setor=setor or "",
            empenho=empenho or "",
            fornecedor=fornecedor or "",
            numero_documento=numero_documento or "",
            data_documento=data_documento or None,
            data_ateste=data_ateste or None,
            dependencia=dependencia or "",
            inventariante=inventariante,
            situacao="localizado"  # default
        )

    # ------------------------------------------------------
    # Ap√≥s inserir os registros, renderiza a tabela atualizada
    # ------------------------------------------------------
    patrimonios = Patrimonio.objects.all().order_by("id")
    paginator = Paginator(patrimonios, 6)
    lista_patrimonios = paginator.page(1)

    context = {
        "pagina": "patrimonio",
        "lista_patrimonios": lista_patrimonios,
        "search_query": "",
        "is_admin": True,
        "user": request.user,
    }

    html = render_to_string("app_inventario/partials/tabela_patrimonios.html", context, request=request)
    response = HttpResponse(html)
    response["HX-Trigger"] = json.dumps({"planilhaAtualizada": True})
    return response


# ==========================================================
# ATUALIZA√á√ÉO R√ÅPIDA DE SITUA√á√ÉO DE PATRIM√îNIO
# ----------------------------------------------------------
# Permite alterar a 'situa√ß√£o' de um patrim√¥nio diretamente
# na tabela via select (HTMX).
# ==========================================================
@login_required
def patrimonio_update_situacao(request, pk):
    # Obt√©m o patrim√¥nio. √â importante garantir que o usu√°rio
    # logado tenha permiss√£o para alterar (regra b√°sica: s√≥ o pr√≥prio inventariante ou admin)
    
    # Tentamos obter o patrim√¥nio, considerando que ele existe
    patrimonio = get_object_or_404(Patrimonio, pk=pk)

    if request.method == 'POST':
        # O HTMX envia o valor do select no corpo da requisi√ß√£o POST
        nova_situacao = request.POST.get('situacao')
        
        if nova_situacao:
            patrimonio.situacao = nova_situacao
            patrimonio.save()
            
            # Ap√≥s salvar, renderizamos o SELECT novamente,
            # for√ßando o HTMX a recarregar apenas o elemento modificado.
            return render(
                request,
                "app_inventario/partials/situacao_select.html", 
                {"p": patrimonio}
            )
        
    # Se a requisi√ß√£o n√£o for POST, apenas retorna o select atual
    return render(
        request,
        "app_inventario/partials/situacao_select.html", 
        {"p": patrimonio}
    )

@login_required
@user_passes_test(presidente_ou_superuser)
def upload_planilha_modal(request):
    """
    Retorna o template que cont√©m o formul√°rio do modal.
    """
    return render(request, "app_inventario/upload_planilha.html")
